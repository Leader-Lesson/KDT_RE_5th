from bs4 import BeautifulSoup
import requests

with open("test.html", "r", encoding="utf-8") as f:
  html_data = f.read()

# soup객체 생성
# soup = BeautifulSoup(html_data, "html.parser") # 내장 파서
soup = BeautifulSoup(html_data, "lxml")
# print(soup)
print(soup.prettify())

# 데이터 선택
# find() - 첫번째 매칭 요소 선택
# 태그를 기준으로 탐색
result = soup.find("h1")
print(result)
print(result.text)
print(result.get_text(strip=True))

# find() - 속성 조건으로 검색 가능
result = soup.find("h1", class_="sub_title")
print(result.text)

# find_all() - 모든 매칭된 요소 선택
result = soup.find_all("h1")
# print(result)

for i in result:
  print(i.text)

# select() - 모든 매칭 요소 선택
# CSS 선택자로 탐색
result = soup.select("ul.items")
print(result)
for i in result:
  print(i.text)

# select_one() - 첫번째 매칭 요소 선택
result = soup.select_one("ul.items")
print(result)

# requests와 함께 사용
# requests : 웹 사이트에서 데이터를 송수신 할 수 있는 라이브러리
# 멜론에서 Top10의 노래 제목의 받아오기
url = "https://www.melon.com/chart/index.htm"
headers = {
  "User-Agent": "Mozilla/5.0"
}

response = requests.get(url, headers=headers)
soup = BeautifulSoup(response.text, "lxml")

songs = soup.select("div.ellipsis.rank01 a")[:10]

for idx, song in enumerate(songs):
  print(f"{idx+1}. {song.text}")

# 실습2. 사용자로부터 입력을 받아 크롤링
import datetime

word = input("검색어를 입력하세요.")

headers = { 
  "User-Agent" : "Mozilla/5.0"
}

response = requests.get(f"https://search.naver.com/search.naver?query={word}",
                        headers=headers)
soup = BeautifulSoup(response.text, "lxml")

news_list = soup.select("a.BxOYkTUC7zH9xrtyOwDx.a2OpSM_aSvFbHwpL_f8N")

print(f"== {datetime.date.today()} 오늘의 축구 뉴스 ==")
for a in news_list:
  print(f"{a.get_text()} / 링크: {a.get("href")}")

# 크롤링한 자료를 엑셀로 저장
import openpyxl

# 엑셀 파일 만들기
wb = openpyxl.Workbook()

# 시트 만들기
ws = wb.create_sheet("codingon")

ws["A1"] = "이름"
ws["B1"] = "영어이름"

ws["A2"] = "코딩오울"
ws["B2"] = "CodingOwl"

wb.save("codingon.xlsx")

import pandas as pd

data = {
  "이름":["이안"],
  "영어이름":["Ian"]
}

df = pd.DataFrame(data)

df.to_excel("condingon2.xlsx", index=False, sheet_name="list")

# 파일 불러오기
wb = openpyxl.load_workbook("codingon.xlsx")

# 시트 선택
ws = wb["codingon"]

ws["A3"] = "김코딩"
ws["B3"] = "CodingKim"

# 여러자료 추가
data = [
  ["안태현", "Ahn"],
  ["이민정", "Lee"],
  ["최하연", "Choi"],
  ["윤태훈", "Yun"],
  ["오왕경", "Oh"],
  ["김진선", "Kim"],
]

for row in data:
  ws.append(row)

wb.save("codingon.xlsx")

# 실습3.
import pandas as pd

res = requests.get("https://finance.naver.com/marketindex/")
soup = BeautifulSoup(res.text, "lxml")
result = soup.select("div.market1 a.head")

data = []
for a in result:
  exchange = a.select_one("span.blind").get_text().split()[1]
  value = a.select_one("span.value")
  # print(exchange, value.get_text())
  data.append([exchange, value.get_text()])

df = pd.DataFrame(data, columns=["통화","환율"])
df.to_excel("exchange.xlsx", index=False)